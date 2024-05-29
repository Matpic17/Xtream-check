import os
import requests
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook, load_workbook

def telecharger_et_lire_m3u(url, username, password):
    m3u_url = f"{url}/get.php?username={username}&password={password}&type=m3u_plus&output=mpegts"
    try:
        response = requests.get(m3u_url, timeout=20)
        if response.status_code == 200:
            return response.text.splitlines()
        else:
            print(f"Erreur lors du téléchargement du fichier M3U. Code: {response.status_code}")
            return None
    except requests.Timeout:
        print(f"Timeout lors du téléchargement du fichier M3U.")
        return None
    except Exception as e:
        print(f"Erreur lors du téléchargement du fichier M3U, erreur : {e}")
        return None

def verifier_disponibilite(films, contenu_m3u):
    languages = ["français", "francais", "french"]
    disponibilite = {film: "" for film in films}
    if contenu_m3u:
        for ligne in contenu_m3u:
            for film in films:
                if film.lower() in ligne.lower():
                    if any(lang in ligne.lower() for lang in languages):
                        disponibilite[film] = "x"
                    elif disponibilite[film] == "" and film.lower() in ligne.lower():
                        disponibilite[film] = "o"
    return disponibilite

def traiter_lien_contenu(url, identifiants, films_a_verifier, ws, wb, fichier_resultat, liens_deja_traites):
    if not url or url in liens_deja_traites:
        return None

    for i, (username, password) in enumerate(identifiants[:5]):  # Essayer jusqu'à 5 combinaisons
        time.sleep(0.1)  # Pause pour limiter le taux de requêtes
        contenu_m3u = telecharger_et_lire_m3u(url, username, password)
        if contenu_m3u:
            resultats = verifier_disponibilite(films_a_verifier, contenu_m3u)

            # Ajouter les résultats au fichier Excel
            row = [url]
            for film in films_a_verifier:
                row.append("x" if resultats[film] == "x" else "o" if resultats[film] == "o" else "")
            ws.append(row)
            wb.save(fichier_resultat)
            return row

    # En cas d'échec de toutes les combinaisons
    if not contenu_m3u:
        row = [url] + ["Erreur de téléchargement"] * len(films_a_verifier)
        ws.append(row)
        wb.save(fichier_resultat)
        return row

def verifier_connexion_serveur(url, username, password):
    m3u_url = f"{url}/get.php?username={username}&password={password}&type=m3u_plus&output=mpegts"
    info_url = f"{url}/player_api.php?username={username}&password={password}&action=get_account_info"
    try:
        response = requests.head(m3u_url, timeout=10)
        if response.status_code == 200:
            response_info = requests.get(info_url, timeout=10)
            if response_info.status_code == 200:
                data = response_info.json()
                if 'user_info' in data and 'exp_date' in data['user_info']:
                    exp_timestamp = int(data['user_info']['exp_date'])
                    exp_date = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(exp_timestamp))
                    return True, exp_date
            return True, None
        return False, None
    except requests.Timeout:
        return False, None
    except Exception as e:
        return False, None

def traiter_lien_connexion(lien, ws, wb, fichier_resultat, liens_deja_traites):
    lien = lien.strip()
    if not lien or lien in liens_deja_traites:
        return None

    parts = lien.split('&')
    url = parts[0].split('?')[0]
    url = url[:-8]
    username = parts[0].split('=')[1]
    password = parts[1].split('=')[1]

    time.sleep(0.1)  # Pause pour limiter le taux de requêtes
    connexion_reussie, exp_date = verifier_connexion_serveur(url, username, password)
    if connexion_reussie:
        row = [lien, "ok", exp_date if exp_date else "N/A"]
    else:
        row = [lien, "Nok", "N/A"]

    ws.append(row)
    wb.save(fichier_resultat)
    return row

def est_fichier_ouvert(fichier):
    """
    Vérifie si un fichier est ouvert par une autre application.
    """
    try:
        with open(fichier, 'a'):
            pass
    except IOError:
        return True
    return False
    

def main():
    mode = input("Entrez le mode (1 pour vérifier le contenu, 2 pour vérifier la connexion) : ").strip()
    if mode not in ["1", "2"]:
        print("Mode invalide.")
        time.sleep(2)
        return
    
    dossier_liens = input("Entrez le chemin du dossier contenant les fichiers texte avec les liens Xtream : ")
    try:
        os.makedirs(dossier_liens, exist_ok=True)
    except OSError as e:
        print(f"Le dossier n'existe pas : {e}")
        return



    films_a_verifier = ["Coraline", "Gravity Falls", "Les enfants du temps", "Freaky Friday", "Le fils à jo", "Bridgerton S03"]



    if mode == "1":
        fichier_resultat = os.path.join(os.path.abspath(os.path.join(dossier_liens, os.pardir)), "mode_1.xlsx")
    else:
        fichier_resultat = os.path.join(os.path.abspath(os.path.join(dossier_liens, os.pardir)), "mode_2.xlsx")

    # Vérifier si le fichier est ouvert
    if os.path.exists(fichier_resultat) and est_fichier_ouvert(fichier_resultat):
        while est_fichier_ouvert(fichier_resultat):
            print(f"Le fichier {fichier_resultat} est actuellement ouvert. Veuillez le fermer et appuyez sur Entrée pour continuer...")
            input()

    if os.path.exists(fichier_resultat):
        wb = load_workbook(fichier_resultat)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        if mode == "1":
            ws.append(["Lien"] + films_a_verifier)
        else:
            ws.append(["Lien", "Statut", "Date d'expiration"])

    liens_deja_traites = set()
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0]:
            lien = row[0].strip()
            liens_deja_traites.add(lien)

    urls_identifiants = {}
    code = []   
    for fichier in os.listdir(dossier_liens):
        if fichier.endswith(".txt"):
            chemin_fichier = os.path.join(dossier_liens, fichier)
            with open(chemin_fichier, 'r', encoding='utf-8') as f:
                liens = f.readlines()
                if mode == "1":
                    for lien in liens:
                        parts = lien.strip().split('&')
                        url = parts[0].split('?')[0]
                        url = url[:-8]
                        username = parts[0].split('=')[1]
                        password = parts[1].split('=')[1]
                        if url not in urls_identifiants:
                            urls_identifiants[url] = []
                        urls_identifiants[url].append((username, password))
                else: #Ajouter tous les liens dans une liste
                    for lien in liens:
                        code.append(lien)
    
    if not urls_identifiants and not code:
        print("Aucun lien trouvé dans le dossier.")
        time.sleep(1.5)
        return        

    with ThreadPoolExecutor(max_workers=50) as executor:
        futures = {}
        if mode == "1":
            futures = {executor.submit(traiter_lien_contenu, url, identifiants, films_a_verifier, ws, wb, fichier_resultat, liens_deja_traites): url for url, identifiants in urls_identifiants.items()}
        else:
            for lien in code:
                futures = {executor.submit(traiter_lien_connexion, lien, ws, wb, fichier_resultat, liens_deja_traites)}

        for future in as_completed(futures):
            result = future.result()
            if result and mode == "1":
                print(f"Lien traité : {result[0]}")

    print(f"Résultats enregistrés dans {fichier_resultat}")
    time.sleep(1.5)

if __name__ == '__main__':
    main()